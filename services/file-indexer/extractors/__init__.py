"""Text extraction dispatch by file extension."""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Optional

logger = logging.getLogger(__name__)


def extract_text(file_path: str) -> Optional[str]:
    """Extract plain text from a file. Returns None for unsupported types."""
    ext = Path(file_path).suffix.lower()

    try:
        if ext in {".txt", ".md", ".csv", ".log", ".json", ".yaml", ".yml",
                    ".xml", ".html", ".css", ".js", ".ts", ".tsx", ".py", ".sh",
                    ".toml", ".ini", ".cfg", ".env"}:
            return _read_plaintext(file_path)
        if ext == ".pdf":
            return _extract_pdf(file_path)
        if ext == ".docx":
            return _extract_docx(file_path)
        if ext == ".xlsx":
            return _extract_xlsx(file_path)
        if ext in {".htm", ".xhtml"}:
            return _extract_html(file_path)
    except Exception as e:
        logger.warning("Extraction failed for %s: %s", file_path, e)
        return None

    return None


def _read_plaintext(path: str) -> str:
    with open(path, "r", encoding="utf-8", errors="replace") as f:
        return f.read()


def _extract_pdf(path: str) -> str:
    import pdfplumber
    pages = []
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            if text:
                pages.append(text)
    return "\n\n".join(pages)


def _extract_docx(path: str) -> str:
    from docx import Document
    doc = Document(path)
    return "\n\n".join(p.text for p in doc.paragraphs if p.text.strip())


def _extract_xlsx(path: str) -> str:
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    parts = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        rows = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            if any(cells):
                rows.append("\t".join(cells))
        if rows:
            parts.append(f"=== Sheet: {sheet} ===\n" + "\n".join(rows))
    wb.close()
    return "\n\n".join(parts)


def _extract_html(path: str) -> str:
    from bs4 import BeautifulSoup
    with open(path, "r", encoding="utf-8", errors="replace") as f:
        soup = BeautifulSoup(f, "html.parser")
    # Strip script/style tags
    for tag in soup(["script", "style"]):
        tag.decompose()
    return soup.get_text(separator="\n", strip=True)
